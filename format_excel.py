import os

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment


def format_excel():

    input_folder = "input"
    output_folder = "output"

    os.makedirs(output_folder, exist_ok=True)

    excel_files = [
        f for f in os.listdir(input_folder)
        if f.endswith(".xlsx")
    ]

    if not excel_files:
        print("No Excel files found.")
        return

    print("\n========== Available Excel Files ==========\n")

    for i, file in enumerate(excel_files, start=1):
        print(f"{i}. {file}")

    while True:

        try:

            choice = int(input("\nChoose file number: "))

            if 1 <= choice <= len(excel_files):
                break

            print("Invalid choice.")

        except ValueError:

            print("Enter a valid integer.")

    filename = excel_files[choice - 1]

    filepath = os.path.join(input_folder, filename)

    workbook = load_workbook(filepath)

    sheet = workbook.active
    # ============================================
    # Header Style
    # ============================================

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ============================================
    # Format Header
    # ============================================

    for cell in sheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ============================================
    # Format Remaining Cells
    # ============================================

    for row in sheet.iter_rows(min_row=2):

        for cell in row:

            cell.border = thin_border

    # ============================================
    # Auto Adjust Column Width
    # ============================================

    for column_cells in sheet.columns:

        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:

                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            except Exception:
                pass

        sheet.column_dimensions[column_letter].width = max_length + 3

    # ============================================
    # Freeze Header
    # ============================================

    sheet.freeze_panes = "A2"

    # ============================================
    # Enable Filters
    # ============================================

    sheet.auto_filter.ref = sheet.dimensions

    # ============================================
    # Save Workbook
    # ============================================

    output_file = os.path.join(
        output_folder,
        f"{os.path.splitext(filename)[0]}_Formatted.xlsx"
    )

    workbook.save(output_file)

    # ============================================
    # Summary
    # ============================================

    print("\n==============================================")
    print("          Excel Formatting Summary")
    print("==============================================")
    print(f"Source File : {filename}")
    print(f"Rows        : {sheet.max_row}")
    print(f"Columns     : {sheet.max_column}")
    print(f"Saved As    : {output_file}")
    print("==============================================")
    print("Excel formatted successfully!")
